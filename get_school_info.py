import requests
import json
import openpyxl
# 高德地图API的URL和你的App Key

APP_KEY = "a007a4876df933705efe6dc17a6a75cd"

#查看某地址的经纬度
def geocode_address(address):
    BASE_URL = "https://restapi.amap.com/v3/geocode/geo"
    params = {
        "key": APP_KEY,
        "address": address,
        "output": "json"  # 输出格式为JSON
    }
    response = requests.get(BASE_URL, params=params)
    data = response.json()  # 解析JSON响应
    if data["status"] == "1" and len(data["count"]) > 0:  # 确保状态码为1且结果数大于0
        return data["geocodes"][0]["location"]  # 返回第一个地理编码结果的位置信息（经度和纬度）
    else:
        print("地址无法解析")
        return None


#查找附近的学校，并且返回学校信息
def find_nearby_schools(lat,lng, radius=50000):
    # BASE_URL = "https://restapi.amap.com/v3/place/text?parameters" #1,没有距离
    BASE_URL = "https://restapi.amap.com/v3/place/around?parameters "  # 2，有距离
    params = {
        "key": APP_KEY,
        "coordtype": "wgs84",  # 坐标类型，这里使用WGS84坐标系
        "location": f"{lat},{lng}",  # 待查询的经纬度坐标
        "output": "json",  # 输出格式为JSON
        "types":"141201",
        "extensions": "all",  # 获取更多扩展信息，包括学校等设施
        "radius": 50000  # 可选参数，查询半径（单位：米）

    }
    response = requests.get(BASE_URL, params=params)
    data = response.json()  # 解析JSON响应
    # print(data)
    # print(data['pois'][5])
    # 提取学校名称和地址信息
    school_names = []  # 存储学校名称的列表
    school_addresses = []  # 存储学校地址的列表
    school_distance = []  # 存储距离的列表
    excel_path = '/Users/lizhenqian/Desktop/school.xlsx' # Excel表格地址，
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook['Sheet1']
    for poi in data['pois']:
        if poi['type']=='科教文化服务;学校;高等院校':  # 确保POI类型以'科教文化服务;学校;高等院校'开头（可根据类型修改）
            school_names.append(poi['name'])  # 提取学校名称并添加到列表中
            school_addresses.append(poi['address'])  # 提取学校地址并添加到列表中
            school_distance.append(poi['distance'])     # 提取相距并添加到列表中
        else:
            print('类型有误，请检查Json！！！！！！！！！！！！！！！！！！！！！！！！！！！！！')


    # 获取学校名称、学校地址、相距多少米，并写入Excel
    worksheet['A1'] = '学校名称'
    worksheet['B1'] = '学校地址'
    worksheet['C1'] = '学校距该地点的距离（米）（一般会比实际少500米）'
    # j决定从第几行开始写
    j = 0
    while j <1:
        print('你想从第几行开始存入Excel：(j最小为1)')
        j = input()
        j = int(j)
        if j <1 :
            print('你输错了！！！重新输入!')
        else:
            break
    for i, value in enumerate(school_names):
        worksheet.cell(row=i+1+j,column = 1,value = value)
        print(value)
    for i, value in enumerate(school_addresses):
        worksheet.cell(row=i+1+j, column = 2, value=value)
    for i, value in enumerate(school_distance):
        worksheet.cell(row=i+1+j, column = 3, value=value)

    print('写入完成!!!')
    workbook.save(excel_path)

address = '淄博市临淄区稷下街道天齐路200甲方正2009南区负一层'
lat_lng = geocode_address(address) #Step1 ：先获取该地址的经纬度
print(f'经度为： {lat_lng[0:10]}, 纬度是：{lat_lng[11:20]}')
lat = lat_lng[11:20] # 纬度---通过geo_test进行查询
lng = lat_lng[0:10]  # 经度---通过geo_test进行查询
school_info = find_nearby_schools(lat,lng, 50000) # Step2 获取学校信息
